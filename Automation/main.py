import openpyxl

file = openpyxl.load_workbook("Book1.xlsx")

product_list = file["Sheet1"]

# here we are making dictionaries for total value and product per supplier
products_per_supplier = {}
total_value_per_supplier = {}
products_less_100 = {}

# here we are iterate excel file rows one by one using for loop
for product_row in range(2, product_list.max_row+1):
    supplier_name = product_list.cell(product_row, 4).value
    price = product_list.cell(product_row, 3).value
    inventory = product_list.cell(product_row, 2).value
    prod_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    if supplier_name in products_per_supplier:
        curr_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = curr_num_products+1
    else:
        print("new added")
        products_per_supplier[supplier_name] = 1

    # calculation total value inventory per supplier
    if supplier_name in total_value_per_supplier:
        curr_supplier_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = curr_supplier_value + (inventory * price)
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # to find inventory product number which have less than 100
    if inventory < 100:
        products_less_100[prod_num] = inventory

    # to add multiplication of inventory and price in excel sheet

    inventory_price.value = inventory * price


print(products_per_supplier)
print(total_value_per_supplier)
print(products_less_100)


file.save("Sheet2.xlsx")

print("EOP")
